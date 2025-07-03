from tkinter import*
import time
from pupil import Pupil
from PIL import Image,ImageTk,ImageGrab
import threading 
from queue import Queue
import cv2
import numpy as np
# import module as m
# import dlib
#import imutils
import datetime
import os
import openpyxl
from tkinter import filedialog, messagebox


def left():
    update()
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    judgebox.insert(0,"left        "+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))
    d.set("left")
def right():
    update()
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    judgebox.insert(0,"right     "+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))
    d.set("right")
def stop():
    update()
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    judgebox.insert(0,"stop      "+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))   
    d.set("stop")
def update():
    global elapsedtime
    global timestr
    global timer
    elapsedtime = time.time()-starttime
    return elapsedtime
def export():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    
    # 取得資料清單
    global data_list  # 使用全局變數


    judge_data = judgebox.get(0, END)
    result_data = resultbox.get(0, END)
    audi_data = audibox.get(0, END)
    history_data = framebox.get(0,END)
    
    # 將兩個 Listbox 的內容寫入 Excel 檔案中
    for i in range(len(judge_data)):
        sheet.cell(row=i+1, column=1, value=judge_data[i])
    for i in range(len(result_data)):
        sheet.cell(row=i+1, column=2, value=result_data[i])
    for i in range(len(audi_data)):
        sheet.cell(row=i+1, column=3, value=audi_data[i])
    for i in range(len(history_data)):
        sheet.cell(row=i+1, column=4, value=history_data[i])
    
     # 將時間戳記和結果以 "時間/結果" 的形式寫入 Excel 檔案中
    # for i, (timestamp, history_data) in enumerate(data_list):
    #     time_result = f"{timestamp}/{history_data}"
    #     sheet.cell(row=i + 1, column=1, value=time_result)

    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    # 檔案儲存路徑不為空時才儲存檔案
    if file_path:
        wb.save(file_path)

    #messagebox.showinfo("Hint","The file has been saved in C:\python_spy")

def Play():
    global running
    global starttime
    if not running:
        starttime = time.time() - elapsedtime
        running = True
        update()
        pass

def record_video():
    global out, recording
    while recording:
        img = ImageGrab.grab()
        
        frame = cv2.cvtColor(np.array(img), cv2.cv2.COLOR_RGB2BGR)

        out.write(frame)

    out.release()

def Start_record():
    global out,  recording, recording_path
    recording=True
    file_name = datetime.datetime.today()
    file_name=str(file_name.strftime("%Y-%m-%d-%H-%M-%S"))
    #file_path = filedialog.asksaveasfilename(defaultextension='.mp4')
    a = os.path.join(recording_path,file_name)
    video_path =os.path.abspath(a)
    out = cv2.VideoWriter(video_path+".mp4", fourcc, 25, (width,high))
    recording_thread = threading.Thread(target=record_video)
    recording_thread.start()
         

def Stop_record():
    global recording, out
    recording = False
    recording_thread.join()
    out.release()

def Reset():
    global elapsedtime
    global starttime
    global running        
    starttime = 0.0
    elapsedtime = 0.0
    if running:
        running = False       

def ControlLeft():
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    audibox.insert(0,"left         "+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))
    

def ControlRight():
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    audibox.insert(0,"right      "+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))
class Calibration(object):
    """
    This class calibrates the pupil detection algorithm by finding the
    best binarization threshold value for the person and the webcam.
    """
    def __init__(self):
        self.nb_frames = 20
        self.thresholds_left = []
        self.thresholds_right = []
    def is_complete(self):
        """Returns true if the calibration is completed"""
        return len(self.thresholds_left) >= self.nb_frames and len(self.thresholds_right) >= self.nb_frames
    def threshold(self, side):
        """Returns the threshold value for the given eye.

        Argument:
            side: Indicates whether it's the left eye (0) or the right eye (1)
        """
        if side == 0:
            return int(sum(self.thresholds_left) / len(self.thresholds_left))
        elif side == 1:
            return int(sum(self.thresholds_right) / len(self.thresholds_right))

    @staticmethod
    def iris_size(frame):
        """Returns the percentage of space that the iris takes up on
        the surface of the eye.

        Argument:
            frame (numpy.ndarray): Binarized iris frame
        """
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        frame = frame[5:-5, 5:-5]
        height, width = frame.shape[:2]
        nb_pixels = height * width
        nb_blacks = nb_pixels - cv2.countNonZero(frame)
        return nb_blacks / nb_pixels

    @staticmethod
    def find_best_threshold(eye_frame):
        """Calculates the optimal threshold to binarize the
        frame for the given eye.

        Argument:
            eye_frame (numpy.ndarray): Frame of the eye to be analyzed
        """
        average_iris_size = 0.48
        trials = {}

        for threshold in range(5, 100, 5):
            iris_frame = Pupil.image_processing(eye_frame, threshold)
            trials[threshold] = Calibration.iris_size(iris_frame)

        best_threshold, iris_size = min(trials.items(), key=(lambda p: abs(p[1] - average_iris_size)))
        return best_threshold

    def evaluate(self, eye_frame, side):
        """Improves calibration by taking into consideration the
        given image.

        Arguments:
            eye_frame (numpy.ndarray): Frame of the eye
            side: Indicates whether it's the left eye (0) or the right eye (1)
        """
        threshold = self.find_best_threshold(eye_frame)

        if side == 0:
            self.thresholds_left.append(threshold)
        elif side == 1:
            self.thresholds_right.append(threshold)
            
def find_pupil(detector, eye):
     # converting frame into Gry image.

    best_threshold = Calibration.find_best_threshold(eye)

    _, img = cv2.threshold(eye, best_threshold, 255, cv2.THRESH_BINARY)#產生灰階圖片進行辨識二值化
    # find the keypoint
    keypoints = detector.detect(img)
    result = None
    if keypoints:
        cv2.drawKeypoints(eye, keypoints, eye, (0, 0, 255),
                          cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)
        sum = 0
        count = [0, 0]
        for k in range(len(keypoints)):
            sum += 1
            for i in range(2):
                count[i] += keypoints[k].pt[i]
        count[0] = count[0]
        count[1] = count[1]

        cv2.circle(eye, (int(count[0]), int(count[1])),
                   radius=3, color=(0, 0, 255), thickness=-1)
        #cv2.imshow('img',img)
        result = count
    return result

data_list = []

def run_analysis():
    global data_list  # 使用全局變數
    Play()
    
    cap = cv2.VideoCapture(r"C:\Users\tseng\Desktop\EyeTracking\2.0\曾X甯_錄影檔 (5)_ 剪_3.0.mp4")
    #cap = cv2.VideoCapture(r"C:\Users\tseng\Downloads\TEST2.mp4")

    
    eye_xml = "python-opencv-detect/haarcascade_eye.xml"
    face_proto_path = os.path.join("face-detection", "deploy.prototxt")
    # face_proto_path = "./caffe/models/bvlc_reference_caffenet/deploy.prototxt"
    face_model_path = os.path.join("face-detection", "res10_300x300_ssd_iter_140000.caffemodel")
    eye_cascade = cv2.CascadeClassifier(eye_xml)
    face_detector = cv2.dnn.readNet(face_proto_path, face_model_path)
    #cap = cv2.VideoCapture(data_file,cv2.CAP_DSHOW)
    frame_counter = 0
    direction = "center"
    detector_params = cv2.SimpleBlobDetector_Params()
    print(detector_params.filterByConvexity)
    print(detector_params.minConvexity)
    detector_params.filterByArea = True
    detector_params.maxArea = 1500
    detector_params.minArea = 100
    # change these to detect pupils
    detector_params.filterByCircularity = True
    detector_params.minCircularity = 0.5
    detector_params.filterByConvexity = True
    detector_params.minConvexity = 0.8
    detector = cv2.SimpleBlobDetector_create(detector_params)
    x=0
    y=0
    z=0

    

    while True:
        frame = cap.read()[1]
        frame = cv2.resize(frame,(560,315))
        #frame = cv2.resize(frame,(0,0),fx=0.7,fy=0.7)  #相機0.7
        frame_counter += 1
        blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300),
                                        (104.0, 177.0, 123.0))
        face_detector.setInput(blob)
        h, w, _ = frame.shape
        detections = face_detector.forward()
        left_eye = None 
        left_pupil = None
        right_eye = None
        right_pupil = None
        
         # 將時間戳記和方向添加到資料清單
        data_list.append([frame_counter, direction])
        
        # Only take the first detected face, and only when confidence > 0.5
        for i in range(0, 1):
            confidence = detections[0, 0, i, 2]
            if confidence < 0.5:
                break
            box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
            (startX, startY, endX, endY) = box.astype("int")
            (startX, startY) = (max(0, startX), max(0, startY))
            (endX, endY) = (min(w - 1, endX), min(h - 1, endY))
            cv2.rectangle(frame, (startX, startY),
                            (endX, endY), (255, 255, 0), 2)
            
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            gray_face = gray_frame[startY:endY, startX:endX]
            face = frame[startY:endY, startX:endX]
            eyes = eye_cascade.detectMultiScale(gray_face, 1.1, 4)

            face_height = endY - startY
            face_width = endX - startX
            sorted_eyes = sorted(eyes, key=lambda x: (x[2], x[3]))
            for (ex, ey, ew, eh) in sorted_eyes:
                if left_eye is not None and right_eye is not None:
                    break
                # Filter if eye is too low
                if ey > face_height * 0.6:
                    continue
                eye_center = ex + ew / 2
                eye = face[ey:ey+eh, ex:ex+ew]
                if eye_center < face_width * 0.5 and left_eye is None:
                    # only keep results with pupil
                    pupil = find_pupil(detector, eye)
                    if pupil:
                        left_eye = (ex, ey, ew, eh)
                        left_pupil = pupil
                elif eye_center > face_width * 0.5 and right_eye is None:
                    # only keep results with pupil
                    pupil = find_pupil(detector, eye)
                    if pupil:
                        right_eye = (ex, ey, ew, eh)
                        right_pupil = pupil
        
        # heuristic for direction detection
        pupil_count = 0
        ratio = 0
        if left_eye is not None:
            pupil_count += 1
            
            ratio += left_pupil[0] / ew
            cv2.rectangle(face, (ex, ey), (ex + ew, ey + eh), (0, 225, 255), 2)
        if right_eye is not None:
            pupil_count += 1
            (ex, ey, ew, eh) = right_eye
            ratio += right_pupil[0] / ew
            cv2.rectangle(face, (
                
                ex, ey), (ex + ew, ey + eh), (0, 225, 255), 2)
        # this can be refined
        if pupil_count > 0:
            ratio /= pupil_count
            if ratio < 0.48:
                direction = "right"
            elif ratio > 0.5:
                direction = "left"
            else:
                direction = "center"
            framebox.insert(0, f"frame: {frame_counter}, direction: {direction}\n")
            filename = f"C:\\Users\\tseng\\Desktop\\EyeTracking\\2.0\\TEST_AUTO_BINARY\\frame_{frame_counter}.jpg"


            cv2.imwrite(filename, frame)
            if "left" in d.get():
                if direction.count('left'):
                    x +=1

                    if x >= 10:
                        resultbox.insert(0, f"baby looks left\n")
                        print(x,z)
                        d.set("stop")
                    '''else:
                        resultbox.insert(0, f"None\n")
                        print(x,z)
                        d.set("stop")'''
                else:
                    z+=1
                
                    if z >= 30:
                        resultbox.insert(0, f"None\n")
                        print(x,z)
                        d.set("stop")
                        
            elif "right" in d.get():        
                if direction.count('right'):
                    y +=1

                    if y >= 10:
                        resultbox.insert(0, f"baby looks right\n")
                        print(y,z)
                        d.set("stop")
                    '''else:
                        resultbox.insert(0, f"None\n")
                        print(y,z)
                        d.set("stop")'''
                else:
                    z+=1
                    if z == 30:
                        resultbox.insert(0, f"None\n")
                        print(y,z)
                        d.set("stop")
            elif "stop" in d.get():        
                x=0
                y=0
                z=0
            
    # Draw the result section
        frame = cv2.cvtColor(frame,cv2.COLOR_BGR2RGB)
        frame = ImageTk.PhotoImage(Image.fromarray(frame))
        A1['image'] = frame
    
        win.update()



win = Tk()
win.title("I-eye")
win.iconbitmap(".\i-eye.ico") #視窗icon #檔名.ico
win.geometry("400x200")
#win.config(background="white")
Label(win,fg="black",text="camera").place()
f1 = LabelFrame(win,bg="black")
f1.place(relx=0.01,rely=0.03)
A1 = Label(f1,bg="black")
A1.pack()

timestr = StringVar()
timestr.set('0:0')
running = False
starttime = 0
elapsedtime = 0.0
timer = None

# folder_name = datetime.datetime.today()
# folder_name=str(folder_name.strftime("%Y-%m-%d-%H-%M-%S"))
# folder_path = filedialog.askdirectory()
# if folder_path:
#     new_folder_path = os.path.join(folder_path,folder_name)
#     os.makedirs(new_folder_path)
#     messagebox.showinfo("Hint",f"The folder has been built in {new_folder_path}")
    
# else:
#     messagebox.showinfo("Hint","The folder has not been built")
# recording_path =os.path.abspath(new_folder_path)
# screen = ImageGrab.grab()
# width, high = screen.size
# fourcc = cv2.VideoWriter_fourcc(*'mp4v')
# out = cv2.VideoWriter(recording_path, fourcc, 25, (width,high))
# recording_thread = None
# recording = True


#img
img1 = PhotoImage(file="./left.png")
img2 = PhotoImage(file="./right.png")
# img3 = PhotoImage(file="./pause.png")
img4 = PhotoImage(file="./record.png")
img5 = PhotoImage(file="./record2.png")

#button
#play_btn = Button(text="開始",bg="skyblue",font=12)
#play_btn.config(command = Play,width=10,height=1)
#play_btn.place(relx=0.1,rely=0.93)
reset_btn = Button(text="Reset", bg = "#81ABFF",font = ("Arial",15,"bold"))
reset_btn.config(command=Reset, width=10,height=2)
reset_btn.place(relx= 0.76, rely=0.86)
left_btn = Button(bg="white")
left_btn.config(command=ControlLeft,image=img1,activebackground = "green")
left_btn.place(relx=0.05,rely=0.78)
right_btn = Button(bg="white")
right_btn.config(command=ControlRight,image=img2,activebackground = "green")
right_btn.place(relx=0.25,rely=0.78)
#stop_btn = Button(bg="white")
#stop_btn.config(command=stop,image=img3,activebackground = "green")
#stop_btn.place(relx=0.15,rely=0.75)
start_btn = Button(command = run_analysis,text = "啟動程式", bg = "#FFFF99",font=("微軟正黑體",15,"bold"))
start_btn.config(width=10,height=1)
start_btn.place(relx=0.145,rely=0.65)
ls_btn=Button(text="left sound",bg="#81ABFF")
ls_btn.config(command=left,width=10,height=1,font =("Arial",15,"bold"))
ls_btn.place(relx=0.03,rely=0.65)
rs_btn=Button(text="right sound",bg="#81ABFF")
rs_btn.config(command=right,width=10,height=1,font=("Arial",15,"bold"))
rs_btn.place(relx=0.256,rely=0.65)
export_btn = Button(bg="#00CC66")
export_btn.config(command=export,text= "Export", bg = "#81ABFF", width=10,height=2,font = ("Arial",15,"bold"))
export_btn.place(relx=0.88,rely=0.86)
record_btn = Button(command=Start_record)
record_btn.config(image=img4,bg="black",activebackground="white")
record_btn.place(relx=0.55,rely=0.85)
end_btn = Button(command=Stop_record)
end_btn.config(image=img5,bg="black",activebackground="white")
end_btn.place(relx=0.65,rely=0.85)


#label
'''RESULT = Label(text="RESULT")
RESULT.config(width=10,height=1,font=12)
RESULT.place(relx=0.15,rely=0.7)'''
TC = Label(text="Testing Control")
TC.config(width=12,height=1,font=("Arial",17,"bold"))
TC.place(relx=0.135,rely=0.58)
sys_result = Label(text="result of system")
sys_result.config(font=("Arial",18,"bold"))
sys_result.place(relx=0.5,rely=0.015)
aud_result = Label(text="result of audiologist")
aud_result.config(font=("Arial",18,"bold"))
aud_result.place(relx=0.77,rely=0.015)
direction = Label(text="direction of sound")
direction.config(font=("Arial",18,"bold"))
direction.place(relx=0.5,rely=0.425)
history = Label(text="history")
history.config(font=("Arial",18,"bold"))
history.place(relx=0.83,rely=0.425)


#menu
menu = Menu(win)

menubar_1 = Menu(menu)                        # 建立第一個選單的子選單，有三個選項
menubar_1.add_command(label="Open")              # 子選單第一個選項
menubar_1.add_command(label="Save",command=export)              # 子選單第二個選項
menubar_1.add_command(label="Exit")              # 子選單第三個選項
menu.add_cascade(label='File', menu=menubar_1) 

win.config(menu=menu)      


#judgebox
frame1 = Frame(win, height=9,width=30)        #direction of sound
frame1.place(relx=0.44,rely=0.48)
scrollbar = Scrollbar(frame1)         # 在頁框中加入捲軸元件
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
a = StringVar()
d = StringVar()
# 在頁框中加入 Listbox 元件，設定 yscrollcommand = scrollbar.set
judgebox = Listbox(frame1,  listvariable=a, height=9, width=27, yscrollcommand = scrollbar.set)
judgebox.config(font=("Arial",16))
judgebox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
scrollbar.config(command = judgebox.yview)  # 設定 scrollbar 的 command = listbox.yview
#resultbox
frame2 = Frame(win, height=9,width=30)        #result of system
frame2.place(relx=0.44,rely=0.07)
scrollbar = Scrollbar(frame2)         # 在頁框中加入捲軸元件
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
b = StringVar()
# 在頁框中加入 Listbox 元件，設定 yscrollcommand = scrollbar.set
resultbox = Listbox(frame2,  listvariable=b, height=9, width=27, yscrollcommand = scrollbar.set)
resultbox.config(font=("Arial",16))
resultbox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
scrollbar.config(command = resultbox.yview)  # 設定 scrollbar 的 command = listbox.yview
#framebox
frame3 = Frame(win, height=9,width=30)
frame3.place(relx=0.72, rely=0.07)
scrollbar = Scrollbar(frame3)         #resulf of audiologist
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
c = StringVar()
audibox = Listbox(frame3,  listvariable=c, height=9, width=27, yscrollcommand = scrollbar.set)
audibox.config(font=("Arial",16))
audibox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
scrollbar.config(command = audibox.yview)  # 設定 scrollbar 的 command = listbox.yview

frame4 = Frame(win, height=9,width=30)          #history
frame4.place(relx=0.72, rely=0.48)
scrollbar = Scrollbar(frame4)         
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
c = StringVar()
framebox = Listbox(frame4,  listvariable=c, height=9, width=27, yscrollcommand = scrollbar.set)
framebox.config(font=("Arial",16))
framebox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
scrollbar.config(command = framebox.yview)  # 設定 scrollbar 的 command = listbox.yview
win.mainloop()



