from tkinter import*
import time
from PIL import Image,ImageTk,ImageGrab
import threading 
from queue import Queue
import numpy as np
import datetime
import os
import openpyxl
from tkinter import filedialog, messagebox
from pupil import Pupil
import cv2
import mediapipe as mp
import numpy as np
import time
import tkinter as tk
import random
import psutil
import gc
#import tracemalloc
#from guppy import hpy
#import objgraph












First_time_click = TRUE
video_path = None
current_label = 0  # 0: 未決定, 1: left, 2: right
prev_label = -1  # 紀錄上一幀人工判別方向 (原本要用來)
prev_ai = -1 
voice_time_hour = 0
voice_time_min = 0
voice_time_sec = 0
voice_judge_data = [] #裡面存放  [ 1 給聲音的時間  2 給聲音的方向   3 人工判斷 對於這次聲音的反應 4 AI判斷 對於這次聲音的反應 5 AI判斷總數 6 AI判斷的失誤數 7 AI的誤判數 ]  的list
voice_judge_data_temp1 = []  # 1 給聲音的時間  2 給聲音的方向   3 人工判斷 對於這次聲音的反應 4 AI判斷 對於這次聲音的反應 temp要去收集這些
voice_judge_data_temp2 = []
voice_judge_data_temp3 = []
temp2_ready = False
temp3_ready = False
framebox = []
audibox = []
resultbox = []
judgebox = []



def enable_ab_buttons():
    rs_btn.config(state=tk.NORMAL)
    ls_btn.config(state=tk.NORMAL)

def enable_cd_buttons():
    right_btn.config(state=tk.NORMAL)
    left_btn.config(state=tk.NORMAL)

def disable_ab_buttons():
    rs_btn.config(state=tk.DISABLED)
    ls_btn.config(state=tk.DISABLED)

def disable_cd_buttons():
    right_btn.config(state=tk.DISABLED)
    left_btn.config(state=tk.DISABLED)

#sound 聲音給予
def left():
    disable_ab_buttons()
    enable_cd_buttons()
    update()
    global voice_time_hour
    global voice_time_min
    global voice_time_sec 
    global judgebox
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime % 60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    voice_time_hour = hours
    voice_time_min = minutes
    voice_time_sec = seconds
    
    judgebox.append("left"+","+str(hours)+":"+str(minutes)+":"+str(seconds)+".")
    
    global voice_judge_data_temp1 
    voice_judge_data_temp1 = []
    voice_judge_data_temp1.append( str(hours)+":"+str(minutes)+":"+str(seconds)+"." ) #加入 1 給聲音的時間
    voice_judge_data_temp1.append( "left" ) #加入 2 給聲音的方向
    global temp2_ready 
    global temp3_ready
    temp2_ready = False
    temp3_ready = False
    
    
    d.set("left")

def right():
    disable_ab_buttons()
    enable_cd_buttons()
    update()
    global voice_time_hour
    global voice_time_min
    global voice_time_sec 
    global judgebox
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime % 60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    voice_time_hour = hours
    voice_time_min = minutes
    voice_time_sec = seconds
    judgebox.append("right"+","+str(hours)+":"+str(minutes)+":"+str(seconds)+".")
    
    global voice_judge_data_temp1 
    voice_judge_data_temp1 = []
    voice_judge_data_temp1.append( str(hours)+":"+str(minutes)+":"+str(seconds)+"." ) #加入 1 給聲音的時間
    voice_judge_data_temp1.append( "right" ) #加入 2 給聲音的方向
    global temp2_ready 
    global temp3_ready
    temp2_ready = False
    temp3_ready = False    
    d.set("right")


def update_display():
    if current_label == 1:
        label.config(text="現在判斷是左")
    elif current_label == 2:
        label.config(text="現在判斷是右")
    elif current_label == 3:
        label.config(text="現在判斷是中")
    else:
        label.config(text="尚未選擇方向")

#聽力師判斷        
def ControlLeft():
    disable_cd_buttons()
    enable_ab_buttons()
    update()
    
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime%60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    global audibox
    global current_label
    current_label = 1  # 設為 left
    audibox.append("left"+","+str(hours)+":"+str(minutes)+":"+str(seconds)+".")


    global voice_judge_data_temp2
    global temp2_ready
    voice_judge_data_temp2.append( "left" ) # 拿到 3 人工判斷 對於這次聲音的反應 需要等拿到4 在進行合併
    temp2_ready = True
    
    update_display()

def ControlRight():  # 
    disable_cd_buttons()
    enable_ab_buttons()
    update()
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime % 60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    global current_label
    global audibox
    current_label = 2  # 設為 right
    audibox.append("right"+","+str(voice_time_hour)+":"+str(voice_time_min)+":"+str(voice_time_sec))
    
    global voice_judge_data_temp2
    global temp2_ready
    voice_judge_data_temp2.append( "right" )  # 拿到 3 人工判斷 對於這次聲音的反應 需要等拿到4 在進行合併
    temp2_ready = True
    
    
    
    update_display()


'''
def ControlForward():
    disable_cd_buttons()
    enable_ab_buttons()
    update()
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    global current_label
    current_label = 3  # 設為 forward
    #audibox.insert(END,"right"+","+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))
    update_display()
'''

def update():
    global elapsedtime
    global timestr
    global timer
    elapsedtime = time.time()-starttime
    return elapsedtime


def import_video(): #選擇要匯入的影片
    global video_path
    file_path = filedialog.askopenfilename(
        title="選擇影片檔案",
        filetypes=[("影片檔案", "*.mp4;*.avi;*.mov;*.mkv"), ("所有檔案", "*.*")]
    )
    if file_path:
        video_path = file_path
        print("選擇的影片路徑:", file_path)  # 這裡可以改成你的處理方式



def export():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'result'
    
    # 取得資料清單
    #global data_list  # 使用全局變數
    global framebox
    global audibox
    global resultbox
    global judgebox
    global voice_judge_data 
    judge_data = judgebox
    result_data = resultbox
    audi_data = audibox
    history_data = framebox


    j = 1



    human_dir_right = 0 
    human_dir_left = 0
    
    all_data = 0
    all_fault = 0
    all_miss = 0
    if len(voice_judge_data) > 0 :
        sheet.cell(row=1, column=j, value="給予聲音的時間")
        sheet.cell(row=1, column=j+1, value="給予聲音的方向")
        sheet.cell(row=1, column=j+2, value="人工對於該次聲音的判斷結果")
        sheet.cell(row=1, column=j+3, value="AI對於該次聲音的判斷結果")
        sheet.cell(row=1, column=j+4, value="AI對於該次判斷的樣本數")
        sheet.cell(row=1, column=j+5, value="AI對於該次判斷的失誤數")
        sheet.cell(row=1, column=j+6, value="AI對於該次判斷的誤判數")
        sheet.cell(row=2, column=j+8, value="AI總樣本數")
        sheet.cell(row=3, column=j+8, value="AI總失誤數")
        sheet.cell(row=4, column=j+8, value="AI總誤判數")
        sheet.cell(row=5, column=j+8, value="AI準確率")
        
        for i in range(len(voice_judge_data)):
            sheet.cell(row=i+2, column=j, value=voice_judge_data[i][0])
            sheet.cell(row=i+2, column=j+1, value=voice_judge_data[i][1])
            sheet.cell(row=i+2, column=j+2, value=voice_judge_data[i][2])
            sheet.cell(row=i+2, column=j+3, value=voice_judge_data[i][3])
            sheet.cell(row=i+2, column=j+4, value=voice_judge_data[i][4])
            sheet.cell(row=i+2, column=j+5, value=voice_judge_data[i][5])
            sheet.cell(row=i+2, column=j+6, value=voice_judge_data[i][6])
            all_data += voice_judge_data[i][4]
            all_fault += voice_judge_data[i][5]
            all_miss += voice_judge_data[i][6]
                

        sheet.cell(row=2, column=j+9, value=all_data)
        sheet.cell(row=3, column=j+9, value=all_fault)
        sheet.cell(row=4, column=j+9, value=all_miss)
        sheet.cell(row=5, column=j+9, value=( all_data - all_fault - all_miss ) / ( all_data - all_miss ) )


    file_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    # 檔案儲存路徑不為空時才儲存檔案
    if file_path:
        wb.save(file_path)

    #messagebox.showinfo("Hint","The file has been saved in C:\python_spy")

def Play():
    global running
    global starttime
    if not running:
        starttime = time.time() - elapsedtime
        running = True
        update()
        pass

def record_video():
    global out, recording
    while recording:
        img = ImageGrab.grab()
        
        frame = cv2.cvtColor(np.array(img), cv2.cv2.COLOR_RGB2BGR)

        out.write(frame)

    out.release()

def stop():
    update()
    global judgebox
    global elapsedtime 
    hours = int(elapsedtime/3600)
    minutes = int(elapsedtime/60)
    seconds = int(elapsedtime-minutes*60)
    hseconds = int((elapsedtime-minutes*60-seconds)*100)
    judgebox.append("stop"+","+str(hours)+":"+str(minutes)+":"+str(seconds)+"."+str(hseconds))   
    d.set("stop")

#嘎嘎嘎嘎嘎
def Recoed_state():
    print("Recoed_state()!!!")
    global recording_or_not
    global record_btn
    if recording_or_not:#如果開始錄影 
        recording_or_not = False #將布林定義為停止錄影，下次按按鈕觸發事件時做判斷
        print("if recording_or_not :")
        print(recording_or_not)
        record_btn.config(image=img5,bg="#CEEDEA",activebackground="#CEEDEA")  #Button 更改成暫停符號||   
        Start_record()#開始錄影
    elif not recording_or_not:#如果停止錄影 
        recording_or_not = True ##將布林定義為開始錄影，下次按按鈕觸發事件時做判斷
        print("elif not recording_or_not:")
        print(recording_or_not)
        record_btn.config(image=img4,bg="#CEEDEA",activebackground="#CEEDEA") #Button 更改成錄影符號 >   
        Stop_record()#停止錄影  

def Start_record():
    global out,  recording, recording_path
    recording=True
    file_name = datetime.datetime.today()
    file_name=str(file_name.strftime("%Y-%m-%d-%H-%M-%S"))
    #file_path = filedialog.asksaveasfilename(defaultextension='.mp4')
    a = os.path.join(recording_path,file_name)
    video_path =os.path.abspath(a)
    out = cv2.VideoWriter(video_path+".mp4", fourcc, 25, (width,high))
    recording_thread = threading.Thread(target=record_video)
    recording_thread.start()
         

def Stop_record():
    global recording, out
    recording = False
    recording_thread.join()
    out.release()

def Reset():
    global elapsedtime
    global starttime
    global running        
    starttime = 0.0
    elapsedtime = 0.0
    if running:
        running = False       


class Calibration(object):
    """
    This class calibrates the pupil detection algorithm by finding the
    best binarization threshold value for the person and the webcam.
    """
    def __init__(self):
        self.nb_frames = 20
        self.thresholds_left = []
        self.thresholds_right = []
    def is_complete(self):
        """Returns true if the calibration is completed"""
        return len(self.thresholds_left) >= self.nb_frames and len(self.thresholds_right) >= self.nb_frames
    def threshold(self, side):
        """Returns the threshold value for the given eye.

        Argument:
            side: Indicates whether it's the left eye (0) or the right eye (1)
        """
        if side == 0:
            return int(sum(self.thresholds_left) / len(self.thresholds_left))
        elif side == 1:
            return int(sum(self.thresholds_right) / len(self.thresholds_right))

    @staticmethod
    def iris_size(frame):
        """Returns the percentage of space that the iris takes up on
        the surface of the eye.

        Argument:
            frame (numpy.ndarray): Binarized iris frame
        """
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        frame = frame[5:-5, 5:-5]
        height, width = frame.shape[:2]
        nb_pixels = height * width
        nb_blacks = nb_pixels - cv2.countNonZero(frame)
        return nb_blacks / nb_pixels

    @staticmethod
    def find_best_threshold(eye_frame):
        """Calculates the optimal threshold to binarize the
        frame for the given eye.

        Argument:
            eye_frame (numpy.ndarray): Frame of the eye to be analyzed
        """
        average_iris_size = 0.48
        trials = {}

        for threshold in range(5, 100, 5):
            iris_frame = Pupil.image_processing(eye_frame, threshold)
            trials[threshold] = Calibration.iris_size(iris_frame)

        best_threshold, iris_size = min(trials.items(), key=(lambda p: abs(p[1] - average_iris_size)))
        return best_threshold

    def evaluate(self, eye_frame, side):
        """Improves calibration by taking into consideration the
        given image.

        Arguments:
            eye_frame (numpy.ndarray): Frame of the eye
            side: Indicates whether it's the left eye (0) or the right eye (1)
        """
        threshold = self.find_best_threshold(eye_frame)

        if side == 0:
            self.thresholds_left.append(threshold)
        elif side == 1:
            self.thresholds_right.append(threshold)




face_mesh = mp.solutions.face_mesh.FaceMesh(min_detection_confidence=0.5, min_tracking_confidence=0.5)

def estimate_head_pose(frame):
    mp_drawing = mp.solutions.drawing_utils
    drawing_spec = mp_drawing.DrawingSpec(thickness=1, circle_radius=1)

    img_h, img_w, img_c = frame.shape
    face_3d = []
    face_2d = []

    # 將色彩空間從BGR轉換為RGB
    image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    # 改善效能
    image.flags.writeable = False

    # 獲取結果
    results = face_mesh.process(image)

    # 改善效能
    image.flags.writeable = True

    # 將色彩空間從RGB轉換為BGR
    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

    text = None  # 初始化 text 為 None
    
    if results.multi_face_landmarks:
        for face_landmarks in results.multi_face_landmarks:
            for idx, lm in enumerate(face_landmarks.landmark):
                if idx == 33 or idx == 263 or idx == 1 or idx == 61 or idx == 291 or idx == 199:
                    if idx == 1:
                        nose_2d = (lm.x * img_w, lm.y * img_h)
                        nose_3d = (lm.x * img_w, lm.y * img_h, lm.z * 3000)

                    x, y = int(lm.x * img_w), int(lm.y * img_h)

                    # 獲取2D坐標
                    face_2d.append([x, y])

                    # 獲取3D坐標
                    face_3d.append([x, y, lm.z])

            # 轉換為NumPy數組
            face_2d = np.array(face_2d, dtype=np.float64)

            # 轉換為NumPy數組
            face_3d = np.array(face_3d, dtype=np.float64)

            # 相機矩陣
            focal_length = 1 * img_w

            cam_matrix = np.array([[focal_length, 0, img_h / 2],
                                   [0, focal_length, img_w / 2],
                                   [0, 0, 1]])

            # 畸變參數
            dist_matrix = np.zeros((4, 1), dtype=np.float64)

            # 解決PnP問題
            success, rot_vec, trans_vec = cv2.solvePnP(face_3d, face_2d, cam_matrix, dist_matrix)

            # 獲取旋轉矩陣
            rmat, jac = cv2.Rodrigues(rot_vec)

            # 獲取角度
            angles, _, _, _, _, _ = cv2.RQDecomp3x3(rmat)

            # 獲取y旋轉度數
            x = angles[0] * 360
            y = angles[1] * 360
            z = angles[2] * 360

            # 判斷使用者的頭部傾斜方向
            if y <= -5:
                text = "Right"
            elif y >= 5:
                text = "Left"
            else:
                text = "Forward"

            #cv2.putText(frame, "y: " + str(np.round(y, 2)), (500, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

    head_pose_info = text

    # 顯式釋放記憶體
    del face_2d
    del face_3d
    return frame, head_pose_info

def find_pupil(detector, eye ):
    # converting frame into Gry image.

    best_threshold = Calibration.find_best_threshold(eye)

    _, img = cv2.threshold(eye, best_threshold, 255, cv2.THRESH_BINARY)#產生灰階圖片進行辨識二值化
    # find the keypoint
    keypoints = detector.detect(img)
    result = None
    if keypoints:
        cv2.drawKeypoints(eye, keypoints, eye, (0, 0, 255),
                          cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)
        sum = 0
        count = [0, 0]
        for k in range(len(keypoints)):
            sum += 1
            for i in range(2):
                count[i] += keypoints[k].pt[i]
        count[0] = count[0]
        count[1] = count[1]

        cv2.circle(eye, (int(count[0]), int(count[1])),
                   radius=3, color=(0, 0, 255), thickness=-1)
        #cv2.imshow('img',img)
        result = count
    return result

data_list = []

def temp_merge():   
    global voice_judge_data
    global voice_judge_data_temp1
    global voice_judge_data_temp2
    global voice_judge_data_temp3
    global temp2_ready
    global temp3_ready 
    temp2_ready = False
    temp3_ready = False
    
    voice_judge_data_temp = [] # 把三個合併 再放入 [ 1 給聲音的時間  2 給聲音的方向   3 人工判斷 對於這次聲音的反應 4 AI判斷 對於這次聲音的反應 ]  
    # 新增 5 AI該次判斷的總數 6 AI該次判斷的失誤數 7 AI該次判斷的誤判數
    voice_judge_data_temp.extend( voice_judge_data_temp1 ) # 1 給聲音的時間  2 給聲音的方向
    voice_judge_data_temp.extend( voice_judge_data_temp2 ) # 3 人工判斷 對於這次聲音的反應
    voice_judge_data_temp.extend( voice_judge_data_temp3 ) # 4 AI判斷 對於這次聲音的反應
    voice_judge_data.append(voice_judge_data_temp) 
    
    print("voice data : " , voice_judge_data )
    
    voice_judge_data_temp1.clear()
    voice_judge_data_temp2.clear()
    voice_judge_data_temp3.clear()
    
    

def run_analysis():
    #Reset()
    global data_list  # 使用全局變數
    global video_path
    global voice_judge_data_temp3
    global temp2_ready
    global temp3_ready
    global framebox
    global audibox
    global resultbox
    
    if not video_path:
        print("請先選擇影片檔案！")
        return
    
    Play()
    framebox.clear()
    audibox.clear()
    resultbox.clear()
    
    cap = cv2.VideoCapture(video_path)
    eye_xml = "python-opencv-detect/haarcascade_eye.xml"
    face_proto_path = os.path.join("face-detection", "deploy.prototxt")
    # face_proto_path = "./caffe/models/bvlc_reference_caffenet/deploy.prototxt"
    face_model_path = os.path.join("face-detection", "res10_300x300_ssd_iter_140000.caffemodel")
    eye_cascade = cv2.CascadeClassifier(eye_xml)
    face_detector = cv2.dnn.readNet(face_proto_path, face_model_path)
    #cap = cv2.VideoCapture(data_file,cv2.CAP_DSHOW)
    frame_counter = 0
    direction = "center"
    detector_params = cv2.SimpleBlobDetector_Params()
    print(detector_params.filterByConvexity)
    print(detector_params.minConvexity)
    detector_params.filterByArea = True
    detector_params.maxArea = 1500
    detector_params.minArea = 100
    # change these to detect pupils
    detector_params.filterByCircularity = True
    detector_params.minCircularity = 0.5
    detector_params.filterByConvexity = True
    detector_params.minConvexity = 0.8
    detector = cv2.SimpleBlobDetector_create(detector_params)
    x=0
    y=0
    z=0
    #tracemalloc.start()
    no_eye = 0
    while True:
        
        
        #tracemalloc.start()
        ret,frame = cap.read()
        if not ret:
            print("無法讀取幀，結束迴圈")
            break
        frame = cv2.resize(frame, (800, 430), interpolation=cv2.INTER_AREA)
        
        #frame = cv2.resize(frame,(0,0),fx=0.7,fy=0.7)  #相機0.7
        frame , head_pose_info= estimate_head_pose(frame)
        frame_counter += 1
        
        blob = None
        blob = cv2.dnn.blobFromImage(frame, 1.0, (300, 300),
                                        (104.0, 177.0, 123.0))#光度
        face_detector.setInput(blob)
        del blob
        
        
        h, w, _ = frame.shape
        detections = face_detector.forward()
        left_eye = None 
        left_pupil = None
        right_eye = None
        right_pupil = None
        
        # 將時間戳記和方向添加到資料清單
        #data_list.append([frame_counter, direction])
        
        # Only take the first detected face, and only when confidence > 0.5
        for i in range(0, 1):
            confidence = detections[0, 0, i, 2]
            if confidence < 0.5:
                break
            box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
            (startX, startY, endX, endY) = box.astype("int")
            (startX, startY) = (max(0, startX), max(0, startY))
            (endX, endY) = (min(w - 1, endX), min(h - 1, endY))
            cv2.rectangle(frame, (startX, startY),
                            (endX, endY), (255, 255, 0), 2)
            
            gray_frame = None
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            gray_face = gray_frame[startY:endY, startX:endX]
            face = frame[startY:endY, startX:endX]
            eyes = eye_cascade.detectMultiScale(gray_face, 1.1, 4)

            face_height = endY - startY
            face_width = endX - startX

            sorted_eyes = sorted(eyes, key=lambda x: (x[2], x[3]))
            for (ex, ey, ew, eh) in sorted_eyes:
                if left_eye is not None and right_eye is not None:
                    y_diff = abs(left_eye[1] - right_eye[1])
                    if y_diff > face_height * 0.3:  # 可根據實際情況調整閾值
                        if left_eye[1] < right_eye[1]:  # 如果左眼比右眼高
                            right_eye = None  # 移除右眼
                        else:  # 如果右眼比左眼高
                            left_eye = None  # 移除左眼
                    else:
                        break
                    
                # Filter if eye is too low
                if ey > face_height * 0.6:
                    continue
                eye_center = ex + ew / 2
                eye = face[ey:ey+eh, ex:ex+ew]
                if eye_center < face_width * 0.5 and left_eye is None:
                    # only keep results with pupil
                    pupil = find_pupil(detector, eye)
                    if pupil:
                        left_eye = (ex, ey, ew, eh)
                        left_pupil = pupil
                elif eye_center > face_width * 0.5 and right_eye is None:
                    # only keep results with pupil
                    pupil = find_pupil(detector, eye)
                    if pupil:
                        right_eye = (ex, ey, ew, eh)
                        right_pupil = pupil

        # heuristic for direction detection
        pupil_count = 0
        ratio = 0
        if left_eye is not None:
            pupil_count += 1
            
            ratio += left_pupil[0] / ew
            cv2.rectangle(face, (ex, ey), (ex + ew, ey + eh), (0, 225, 255), 2)
        if right_eye is not None:
            pupil_count += 1
            (ex, ey, ew, eh) = right_eye
            ratio += right_pupil[0] / ew
            cv2.rectangle(face, (ex, ey), (ex + ew, ey + eh), (0, 225, 255), 2)
            
        # this can be refined
        #print(f"pupil_count : {pupil_count}")
        #print(f"head_pose_info : {head_pose_info}")
        direction = "not"
        
        global prev_ai 
        current_ai = -1
        
        
        #print(f"prev_ai : {prev_ai}")
        
        if pupil_count > 0:
            ratio /= pupil_count
            #print(f"ratio : {ratio}")
            if ratio < 0.48:
                direction = "right"
            elif ratio > 0.50:
                direction = "left"
            else:
                direction = "forward"
            update()
            hours = int(elapsedtime/3600)
            minutes = int(elapsedtime/60)
            seconds = int(elapsedtime-minutes*60)
            hseconds = int((elapsedtime-minutes*60-seconds)*100)
            
            framebox.append("time,"+f"{hours}:{minutes}:{seconds}.{hseconds}"+f",direction,{direction},head_pose,{head_pose_info}")
            #framebox.insert(END, "time,"+f"{hours}:{minutes}:{seconds}.{hseconds}"+f",direction,{direction},head_pose,{head_pose_info}")
            
            
            
            if current_label == 1:
                label_text = "left"
            elif current_label == 2:
                label_text = "right"
            elif current_label == 3:
                label_text = "forward"
            else:
                label_text = "no judge"

            global prev_label 
            
            
            
            #if prev_label != current_label : #方向不同才紀錄
            #    audibox.insert(END, f"{label_text},{hours}:{minutes}:{seconds}.{hseconds}") ###############################################################################
            
            prev_label = current_label  # 更新上一幀方向
 

            #print(f"direction: {direction}, head_pose_info: {head_pose_info}")


            # 檢查方向和頭部姿勢信息是否符合您的條件
            if direction == "right" and head_pose_info == "Right":
                # 在圖像上添加 "RIGHT" 文本
                cv2.putText(frame, "RIGHT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                #print( "畫面輸出 : Right1" )
                current_ai = 2
            elif direction == "left" and head_pose_info == "Left":
                # 在圖像上添加 "LEFT" 文本
                cv2.putText(frame, "LEFT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                #print( "畫面輸出 : LEFT1" )
                current_ai = 1
            elif direction == "forward" and head_pose_info == "Forward":
                # 在圖像上添加 "LEFT" 文本
                cv2.putText(frame, "Forward", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                #print( "畫面輸出 : Forward1" )
                current_ai = 3
            else :
                if head_pose_info == "Right":
                    cv2.putText(frame, "RIGHT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : Right2" )
                elif head_pose_info == "Left":
                    cv2.putText(frame, "LEFT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : LEFT2" )
                else :
                    cv2.putText(frame, "Forward", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : Forward2" )
               

                


            #filename = "C:\\Users\\tseng\\Desktop\\EyeTracking\\20221205\\FRAME_2\\frame_{}.jpg".format(frame_counter)

            #cv2.imwrite(filename, frame)


            # d 會根據聲音給予的方向 假設現在聲音是左邊 就會在給予聲音後 一定時間內 判斷聲音方向
            # 根據該幀 判斷的方向給予累計 例如左邊如果在一定時間內 出現的次數最多 那麼代表嬰兒看向左邊 
            if "left" in d.get():
                if direction.count('left') or head_pose_info == "Left" :
                    x +=1  # X 為向前看的指標 代表嬰兒向左看
                    if x >= 10:
                        resultbox.append(f"baby looks,left\n")
                        print(x,z)
                        d.set("stop")
                        if not temp3_ready :
                            voice_judge_data_temp3.append( "left" )  # 拿到 4 人工判斷 對於這次聲音的反應 需要等拿到3 在進行合併
                            voice_judge_data_temp3.append( x + y + z + no_eye ) # 5 AI的該次判斷的總數
                            voice_judge_data_temp3.append( y + z ) # 6 其他兩個方向的數量為失誤
                            voice_judge_data_temp3.append( no_eye ) # 沒找到眼睛 是誤判
                            temp3_ready = True
                else:
                    z+=1  # z 為向前看的指標 代表嬰兒沒對聲音反應
                    if z >= 30:
                        resultbox.append(f"baby looks,forward\n")
                        print(x,z)
                        d.set("stop")
                        if not temp3_ready :
                            voice_judge_data_temp3.append( "forward" )  # 拿到 4 人工判斷 對於這次聲音的反應 需要等拿到3 在進行合併
                            voice_judge_data_temp3.append( x + y + z + no_eye ) # 5 AI的該次判斷的總數
                            voice_judge_data_temp3.append( x + y ) # 6 其他兩個方向的數量為失誤
                            voice_judge_data_temp3.append( no_eye ) # 沒找到眼睛 是誤判
                            temp3_ready = True
            elif "right" in d.get():        
                if direction.count('right') or head_pose_info == "Right" :
                    y +=1  # y 為向右看的指標 代表嬰兒向右看
                    if y >= 10:
                        resultbox.append(f"baby looks,right\n")
                        print(y,z)
                        d.set("stop")
                        if not temp3_ready :
                            voice_judge_data_temp3.append( "right" )  # 拿到 4 人工判斷 對於這次聲音的反應 需要等拿到3 在進行合併
                            voice_judge_data_temp3.append( x + y + z + no_eye ) # 5 AI的該次判斷的總數
                            voice_judge_data_temp3.append( x + z  ) # 其他兩個方向的數量為失誤
                            voice_judge_data_temp3.append( no_eye ) # 沒找到眼睛 是誤判
                            temp3_ready = True
                else:
                    z+=1 # z 為向前看的指標 代表嬰兒沒對聲音反應
                    if z >= 30:
                        resultbox.append(f"baby looks,forward\n")
                        print(y,z)
                        d.set("stop")
                        if not temp3_ready :
                            voice_judge_data_temp3.append( "forward" )  # 拿到 4 人工判斷 對於這次聲音的反應 需要等拿到3 在進行合併
                            voice_judge_data_temp3.append( x + y + z + no_eye ) # 5 AI的該次判斷的總數
                            voice_judge_data_temp3.append( x + y ) # 其他兩個方向的數量為失誤
                            voice_judge_data_temp3.append( no_eye ) # 沒找到眼睛 是誤判
                            temp3_ready = True
            elif "stop" in d.get():        
                x=0
                y=0
                z=0
                no_eye = 0
            
                
                
                
        elif pupil_count  == 0 :
            if "stop" in d.get() :
                no_eye = 0
            else :
                no_eye += 1
            
            if head_pose_info == "Right":
                cv2.putText(frame, "RIGHT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                #print( "畫面輸出 : Right4" )
            elif head_pose_info == "Left" :
                cv2.putText(frame, "LEFT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                #print( "畫面輸出 : LEFT4" )
            else :
                if prev_ai == 1:
                    cv2.putText(frame, "LEFT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : LEFT5" )
                elif prev_ai == 2:
                    cv2.putText(frame, "RIGHT", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : RIGHT5" )
                elif prev_ai == 3:
                    cv2.putText(frame, "Forward", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
                    #print( "畫面輸出 : Forward5" )


        #print(f"current_ai : {current_ai}")

        if current_ai != -1 :
            prev_ai = current_ai
            
            
        if temp2_ready and temp3_ready :
            temp_merge()
            
            
            
    # Draw the result section

        frame = cv2.cvtColor(frame,cv2.COLOR_BGR2RGB)
        if hasattr(A1, 'image') and A1.image is not None:
            A1.image = None  # 解除 Label 的影像綁定
            
        frame = ImageTk.PhotoImage(Image.fromarray(frame))
        A1['image'] = frame
        win.update()
        
        del frame
        '''
        pid = os.getpid()  # 取得當前程式的進程 ID
        process = psutil.Process(pid)
        mem_info = process.memory_info()
        print(f"目前記憶體使用量: {mem_info.rss / 1024 ** 2:.2f} MB")  # 轉換成 MB
        '''
        

        
        
        


#tracemalloc.start()
# 創建主視窗
win = Tk()
win.title("I-eye")
win.iconbitmap(r"./i-eye.ico") #視窗icon #檔名.ico
win.geometry("1200x650")
#win.config(background="#BFE6E3")


timestr = StringVar()
timestr.set('0:0')
running = False
starttime = 0
elapsedtime = 0.0
timer = None
Reset()


#menu
menu = Menu(win)

menubar_1 = Menu(menu)                        # 建立第一個選單的子選單，有三個選項
menubar_1.add_command(label="Open")              # 子選單第一個選項
menubar_1.add_command(label="Save",command=export)              # 子選單第二個選項
menubar_1.add_command(label="Exit")              # 子選單第三個選項
menu.add_cascade(label='File', menu=menubar_1) 

win.config(menu=menu)      


#judgebox
frame1 = Frame(win, height=9,width=30)        #direction of sound
frame1.place(relx=0.44,rely=0.48)
scrollbar = Scrollbar(frame1)         # 在頁框中加入捲軸元件
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
a = StringVar()
d = StringVar()
# 在頁框中加入 Listbox 元件，設定 yscrollcommand = scrollbar.set
#judgebox = Listbox(frame1,  listvariable=a, height=9, width=27, yscrollcommand = scrollbar.set)
#judgebox.config(font=("Arial",16))
#judgebox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
#scrollbar.config(command = judgebox.yview) # 設定 scrollbar 的 command = listbox.yview
#resultbox
frame2 = Frame(win, height=9,width=30)        #result of system
frame2.place(relx=0.44,rely=0.07) 
scrollbar = Scrollbar(frame2)         # 在頁框中加入捲軸元件
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
b = StringVar()
# 在頁框中加入 Listbox 元件，設定 yscrollcommand = scrollbar.set
#resultbox = Listbox(frame2,  listvariable=b, height=9, width=27, yscrollcommand = scrollbar.set)
#resultbox.config(font=("Arial",16))
#resultbox.pack(side='left', fill='y')    # 設定 Listbox 的位置以及填滿方式
#scrollbar.config(command = resultbox.yview)  # 設定 scrollbar 的 command = listbox.yview
#framebox
frame3 = Frame(win, height=9,width=30)
frame3.place(relx=0.72, rely=0.07)
scrollbar = Scrollbar(frame3)         #resulf of audiologist
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
c = StringVar()


 #history 
frame4 = Frame(win, height=1,width=1)         
frame4.place(relx=0.72, rely=0.48)
scrollbar = Scrollbar(frame4)         
scrollbar.pack(side='right', fill='y')  # 設定捲軸的位置以及填滿方式
c = StringVar()




#img
img0 = ImageTk.PhotoImage(Image.open(r"./background3.png"))
img1 = ImageTk.PhotoImage(Image.open(r"./Left_.png"))
#img2 = PhotoImage(file="./Right_.png")
img2 = ImageTk.PhotoImage(Image.open(r"./Right_.png"))

# img3 = PhotoImage(file="./pause.png")
img4 = ImageTk.PhotoImage(Image.open(r"./Record_.png"))
img5 = ImageTk.PhotoImage(Image.open(r"./stop.png"))
img6 = ImageTk.PhotoImage(Image.open(r"./CHF.png"))
img7 = ImageTk.PhotoImage(Image.open(r"./Audio_Judg.png"))
img8 = ImageTk.PhotoImage(Image.open(r"./D_S.png"))
img9 = ImageTk.PhotoImage(Image.open(r"./AI_R.png"))
img10 = ImageTk.PhotoImage(Image.open(r"./sq.png"))
img11 = ImageTk.PhotoImage(Image.open(r"./export.png"))
img12 = ImageTk.PhotoImage(Image.open(r"./start.png"))
img13 = ImageTk.PhotoImage(Image.open(r"./fin.png"))
img14 = ImageTk.PhotoImage(Image.open(r"./mid.png"))

#按鈕布林
recording_or_not = True



label = tk.Label(win, image=img0)
label.pack(fill=tk.BOTH, expand=tk.YES)


label = tk.Label(win, image=img10,bd=0)
label.place(relx=0.2,rely=0.005)#Right
#eye_aud




# 顯示當前選擇的方向
label = tk.Label(win, text="未選擇方向", font=("Arial", 24) , bg="white", compound="top")
label.place(x=10, y=250)  # 固定位置 (10,10)



#聽力師判斷按鍵
left_btn = Button(bg="#EDF8F7",bd=0)
left_btn.config(command=ControlLeft,image=img1,activebackground = "#EDF8F7", state=tk.DISABLED)
left_btn.place(relx=0.63,rely=0.855)
right_btn = Button(bg="#EDF8F7",bd=0)
right_btn.config(command=ControlRight,image=img2,activebackground = "#EDF8F7", state=tk.DISABLED)
right_btn.place(relx=0.77,rely=0.855)

'''
right_btn = Button(bg="#EDF8F7",bd=0)
right_btn.config(command=ControlForward,image=img14,activebackground = "#EDF8F7", state=tk.NORMAL)
right_btn.place(relx=0.705,rely=0.860)
'''


#ControlForward
#sound button
ls_btn= Button(bg="#EDF8F7",bd=0)
ls_btn.config(command=left,image=img1,activebackground = "#EDF8F7",  state=tk.NORMAL)
ls_btn.place(relx=0.19,rely=0.855)
rs_btn=Button(bg="#EDF8F7",bd=0)
rs_btn.config(command=right,image=img2,activebackground = "#EDF8F7",  state=tk.NORMAL)
rs_btn.place(relx=0.35,rely=0.855)



start_btn = tk.Button(win, command=run_analysis, text="START", bd=0, bg="#FFEEEE", font=("微軟正黑體", 12, "bold"), fg="red")
start_btn.config(width=5,height=1,activebackground = "#FFEEEE")
start_btn.place(relx=0.482,rely=0.878)

#匯出按鍵
export_btn = Button(bd=0)
export_btn.config(command=export,image= img11, bg = "#CEEDEA",activebackground="#CEEDEA")
export_btn.place(relx=0.25,rely=0.012)

#匯入影片按鍵
record_btn = Button(bd=0,command=import_video)
record_btn.config(image=img4,bg="#CEEDEA",activebackground="#CEEDEA")
record_btn.place(relx=0.215,rely=0.0124)
  


#攝影機
Label(win,fg="black",text="camera").place()
f1 = LabelFrame(win,bg="black")
f1.place(relx=0.18,rely=0.068)#調整攝影機位置
A1 = Label(f1,bg="black")
A1.pack()






'''    
def check_memory():
    snapshot = tracemalloc.take_snapshot()

    # 顯示最佔用記憶體的程式碼行
    top_stats = snapshot.statistics('lineno')

    print("[ 記憶體使用量排行 ]")
    for stat in top_stats[:10]:
        print(stat)
        
    print("[ HPP ]")
    hp = hpy()
    print(hp.heap())  # 顯示目前記憶體使用情況
    print("[ 顯示 10 種最多的物件 ]")
    objgraph.show_most_common_types(limit=10)  # 顯示 10 種最多的物件
    
    win.after(5000, check_memory)  # 每 5 秒更新一次
    
'''
    
def collect():
    #print("collect")
    gc.collect()   
    
    '''
    pid = os.getpid()  # 取得當前程式的進程 ID
    process = psutil.Process(pid)
    mem_info = process.memory_info()
    print(f"目前記憶體使用量: {mem_info.rss / 1024 ** 2:.2f} MB")  # 轉換成 MB
    
    '''
    win.after(20000, collect)  # 每 5 秒gc一次 

#check_memory()
collect()
win.mainloop()
enable_ab_buttons()


